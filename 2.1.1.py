import csv
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side

#Класс вакансии
class Vacancy:
    #словарь для перевода валют в рубли
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                       "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }

    # Конструктор для объектов класса Vacancy (вакансия)
    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salaryTo = int(float(vacancy['salary_to']))
        self.salaryFrom = int(float(vacancy['salary_from']))
        self.salaryCurrency = vacancy['salary_currency']
        self.salaryAverage = ((self.salaryFrom + self.salaryTo) * self.currency_to_rub[self.salaryCurrency]) / 2
        self.areaName = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

#Класс данных, который отвечает за чтение и подготовку данных из CSV-файла (универсальный парсер CSV)
class DataSet:
    # Конструктор для объектов класса DataSet(данные)
    def __init__(self, file_name, vacancy_name):
        self.fileName = file_name
        self.vacancyName = vacancy_name

    #метод заносит данные в словарь
    @staticmethod
    def increment(dict, key, quantity):
        #если по такому ключу ещё нет значения, то задаём его
        if key in dict:
            dict[key] += quantity
        else:
            dict[key] = quantity

    #метод находит среднее по значениям в заданном словаре
    @staticmethod
    def average(dic):
        dictionary = {}
        # в цикле проходим по всем значениям словаря и находим среднее
        for key, values in dic.items():
            dictionary[key] = int(sum(values)) / int(len(values))
        return dictionary #возвращаем словарь средних значений

    # метод читает заданный файл
    def csv_reader(self):
        # читаем файл
        with open(self.fileName, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            #первую строку записываем отдельно, так как в ней хранятся названия атрибутов
            first = next(reader)
            # проверяем, чтобы были введены все данные
            for i in reader:
                if len(i) == len(first) and '' not in i:
                    yield dict(zip(first, i))

    #метод по данным создаёт статистики
    def get_statistic(self):
        #создаём необходимые словари
        salary = {}
        salaryVacancy = {}
        salaryCity = {}
        count = 0 #вспомагательная переменная-счётчик
        reader = self.csv_reader()
        # в цикле проходимся по всей прочитанной из файла информации по вакансиям
        for vacancyDict in reader:
            vacancy = Vacancy(vacancyDict) #задаём рассматриваемую вакансию
            self.increment(salary, vacancy.year, [vacancy.salaryAverage])
            if vacancy.name.find(self.vacancyName) != -1:
                self.increment(salaryVacancy, vacancy.year, [vacancy.salaryAverage])
            self.increment(salaryCity, vacancy.areaName, [vacancy.salaryAverage])
            count += 1 #считаем количество вакансий
        #начинаем заполнять данные
        number = dict([(key, len(value)) for key, value in salary.items()])
        numbersName = dict([(key, len(value)) for key, value in salaryVacancy.items()])
        if not salaryVacancy:
            salaryVacancy = dict([(key, [0]) for key, value in salary.items()])
            numbersName = dict([(key, 0) for key, value in number.items()])
        statistics = self.average(salary)
        statistics2 = self.average(salaryVacancy)
        statistics3 = self.average(salaryCity)
        statistics4 = {}
        #заполняем данные по ЗП
        for year, salaries in salaryCity.items():
            statistics4[year] = round(len(salaries) / count, 4)
        statistics4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in statistics4.items()]))
        statistics4.sort(key=lambda a: a[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda a: a[0] in list(statistics4.keys()), [(key, value) for key, value in statistics3.items()]))
        statistics3.sort(key=lambda a: a[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])
        return statistics, number, statistics2, numbersName, statistics3, statistics5 #возвращаем полученные данные

    #метод выводит все полученные статистики
    @staticmethod
    def print_statistic(statistics, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))

class Report:
    # Конструктор для объектов класса Report(отчёт)
    def __init__(self, vacancyName, statistics, statistics2, statistics3, statistics4, statistics5, statistics6):
        self.workbook = Workbook()
        self.vacancyName = vacancyName
        self.statistics = statistics
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    #метод формирует excel файл
    def get_excel(self):
        c = self.workbook.active
        #задаём название таблицы и атрибуиов
        c.title = 'Статистика по годам'
        c.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
                  'Количество вакансий - ' + self.vacancyName])
        for year in self.statistics.keys():
            c.append([year, self.statistics[year], self.statistics3[year], self.statistics2[year],
                      self.statistics4[year]])
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        #в цикле находим необходимое количесвто столбцов
        columnWidth = []
        for x in data:
            for i, cell in enumerate(x):
                if len(columnWidth) > i:
                    if len(cell) > columnWidth[i]:
                        columnWidth[i] = len(cell)
                else:
                    columnWidth += [len(cell)]
        for i, columnWidth in enumerate(columnWidth, 1):
            #устанавливаем ширину столбцов для вместимости самой длинной строки в столбце
            c.column_dimensions[get_column_letter(i)].width = 2 + columnWidth
        newData = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city, value), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()):
            newData.append([city, value, '', city2, value2])
        d = self.workbook.create_sheet('Статистика по городам')
        for x in newData:
            d.append(x)
        # в цикле находим необходимое количесвто столбцов
        columnWidth = []
        for x in newData:
            for i, cellule in enumerate(x):
                cellule = str(cellule)
                if len(columnWidth) > i:
                    if len(cellule) > columnWidth[i]:
                        columnWidth[i] = len(cellule)
                else:
                    columnWidth += [len(cellule)]
        for i, columnWidth in enumerate(columnWidth, 1):
            #устанавливаем ширину столбцов для вместимости самой длинной строки в столбце
            d.column_dimensions[get_column_letter(i)].width = columnWidth + 2
        fontBold = Font(bold=True)
        #заполняем таблицу данными
        for column in 'ABCDE':
            c[column + '1'].font = fontBold
            d[column + '1'].font = fontBold
        for index, _ in enumerate(self.statistics5):
            d['E' + str(2 + index)].number_format = '0.00%'
        fontThin = Side(border_style='thin', color='00000000')
        for x in range(len(newData)):
            for column in 'ABDE':
                d[column + str(1 + x)].border = Border(left=fontThin, bottom=fontThin, right=fontThin, top=fontThin)
        self.statistics[1] = 1
        for x, _ in enumerate(self.statistics):
            for column in 'ABCDE':
                c[column + str(1 + x)].border = Border(left=fontThin, bottom=fontThin, right=fontThin, top=fontThin)
        self.workbook.save('report.xlsx')

#Класс отвечает за обработку параметров вводимых пользователем: фильтры, сортировка, диапазон вывода, требуемые столбцы, а также за вывод таблицы в excel
class InputConnect:
    # Конструктор для объектов класса InputConnect (входное соединение)
    def __init__(self):
        self.fileName = input('Напишите название файла: ')
        self.vacancyName = input('Введите название профессии: ')
        a = DataSet(self.fileName, self.vacancyName)
        statistics, statistics2, statistics3, statistics4, statistics5, statistics6 = a.get_statistic()
        a.print_statistic(statistics, statistics2, statistics3, statistics4, statistics5, statistics6)
        b = Report(self.vacancyName, statistics, statistics2, statistics3, statistics4, statistics5, statistics6)
        b.get_excel()

if __name__ == '__main__':
    InputConnect()
