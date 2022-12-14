#импортируем все необходимые библиотеки
import csv
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import Border, Font, Side
import matplotlib.pyplot as plt
import numpy as np
import pathlib
import pdfkit
from jinja2 import Environment, FileSystemLoader

class Vacancy:
    """
        Класс для представления вакансии
        Attributes:
            name (str): Название вакансии
            salaryTo (int): Нижняя граница зарплат
            salaryFrom (int): Верхняя граница зарплат
            salaryCurrency (str): Идентификатор валюты оклада
            salaryAverage (float): Средняя зарплата
            areaName (str): Название региона
            year (int): Год публикации вакансии
            currency_to_rub (dict): Словарь курса валют
    """
    currency_to_rub = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                       "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055, }

    def __init__(self, vacancy):
        """Инициализирует объект Vacancy
                Args:
                    vacancy (dict): Информация об 1 вакансии, полученная в виде словаря из файла
        """
        self.name = vacancy['name']
        self.salaryTo = int(float(vacancy['salary_to']))
        self.salaryFrom = int(float(vacancy['salary_from']))
        self.salaryCurrency = vacancy['salary_currency']
        self.salaryAverage = ((self.salaryFrom + self.salaryTo) * self.currency_to_rub[self.salaryCurrency]) / 2
        self.areaName = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

class DataSet:
    """Класс отвечает за чтение и подготовку данных из CSV-файла
        Attributes:
            fileName (str): Название CSV-файла
            vacancyName (str): Название профессии для статистики
    """

    def __init__(self, file_name, vacancy_name):
        """Инициализирует объект DataSet
                Args:
                    fileName (str): Имя файла
                    vacancyName (str): Название профессии, по которой создаётся статистика
        """
        self.fileName = file_name
        self.vacancyName = vacancy_name

    @staticmethod
    def increment(dict, key, quantity):
        """Заносит данные в словарь
                Args:
                    dict (dict): Словарь зарплат
                    key (int): Год публикации вакансии
                    quantity (list): Средняя зарплата
        """
        if key in dict:
            dict[key] += quantity
        else:
            dict[key] = quantity

    @staticmethod
    def average(dic):
        """Считает средние значения по годам
                Returns: Словарь средних величин по годам
                Args:
                    dict: словарь по годам со статистической информацией
            """
        dictionary = {}
        # в цикле проходим по всем значениям словаря и находим среднее
        for key, values in dic.items():
            dictionary[key] = int(sum(values)) / int(len(values))
        return dictionary #возвращаем словарь средних значений

    """Парсит файл и формирует из вакансий словари
            Returns:
                dict: Словари с информацией по вакансиям
                генератор при помощи yield
    """
    def csv_reader(self):
        # читаем файл
        with open(self.fileName, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            #первую строку записываем отдельно, так как в ней хранятся названия атрибутов
            first = next(reader)
            # проверяем, чтобы были введены все данные
            for i in reader:
                if len(i) == len(first) and '' not in i:
                    yield dict(zip(first, i))

    """Формирует статистику
            Returns:
                tuple: Картеж из словарей со статистикой по городам и годам
    """
    def get_statistic(self):
        #создаём необходимые словари
        salary = {}
        salaryVacancy = {}
        salaryCity = {}
        count = 0 #вспомагательная переменная-счётчик
        reader = self.csv_reader()
        # в цикле проходимся по всей прочитанной из файла информации по вакансиям
        for vacancyDict in reader:
            vacancy = Vacancy(vacancyDict) #задаём рассматриваемую вакансию
            self.increment(salary, vacancy.year, [vacancy.salaryAverage])
            if vacancy.name.find(self.vacancyName) != -1:
                self.increment(salaryVacancy, vacancy.year, [vacancy.salaryAverage])
            self.increment(salaryCity, vacancy.areaName, [vacancy.salaryAverage])
            count += 1 #считаем количество вакансий
        #начинаем заполнять данные
        number = dict([(key, len(value)) for key, value in salary.items()])
        numbersName = dict([(key, len(value)) for key, value in salaryVacancy.items()])
        if not salaryVacancy:
            salaryVacancy = dict([(key, [0]) for key, value in salary.items()])
            numbersName = dict([(key, 0) for key, value in number.items()])
        statistics = self.average(salary)
        statistics2 = self.average(salaryVacancy)
        statistics3 = self.average(salaryCity)
        statistics4 = {}
        #заполняем данные по ЗП
        for year, salaries in salaryCity.items():
            statistics4[year] = round(len(salaries) / count, 4)
        statistics4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in statistics4.items()]))
        statistics4.sort(key=lambda a: a[-1], reverse=True)
        statistics5 = statistics4.copy()
        statistics4 = dict(statistics4)
        statistics3 = list(filter(lambda a: a[0] in list(statistics4.keys()), [(key, value) for key, value in statistics3.items()]))
        statistics3.sort(key=lambda a: a[-1], reverse=True)
        statistics3 = dict(statistics3[:10])
        statistics5 = dict(statistics5[:10])
        return statistics, number, statistics2, numbersName, statistics3, statistics5 #возвращаем полученные данные

    """Выводит статистику
            Args:
                statistics (dict): Динамика уровня зарплат по годам
                statistics2 (dict): Динамика количества вакансий по годам
                statistics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
                statistics4 (dict): Динамика количества вакансий по годам для выбранной профессии
                statistics5 (dict): Уровень зарплат по городам (в порядке убывания)
                statistics6 (dict): Доля вакансий по городам (в порядке убывания)
    """
    @staticmethod
    def print_statistic(statistics, statistics2, statistics3, statistics4, statistics5, statistics6):
        print('Динамика уровня зарплат по годам: {0}'.format(statistics))
        print('Динамика количества вакансий по годам: {0}'.format(statistics2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(statistics3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(statistics4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(statistics5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(statistics6))

#Класс отчёта (excel, графики, pdf)
class Report:
    """Класс для формирования файлов: exel, pdf, изображение диаграмм
        Attributes:
            workbook (Workbook): экземпляр книги для эксель файла
            vacancyName (str): Название профессии
            statistics (dict): Динамика уровня зарплат по годам
            statistics2 (dict): Динамика количества вакансий по годам
            statistics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
            statistics4 (dict): Динамика количества вакансий по годам для выбранной профессии
            statistics5 (dict): Уровень зарплат по городам (в порядке убывания)
            statistics6 (dict): Доля вакансий по городам (в порядке убывания)
        """
    def __init__(self, vacancyName, statistics, statistics2, statistics3, statistics4, statistics5, statistics6):
        """Инициализирует объект Report
                Args:
                    vacancyName (str): Название профессии
                    statistics (dict): Динамика уровня зарплат по годам
                    statistics2 (dict): Динамика количества вакансий по годам
                    statistics3 (dict): Динамика уровня зарплат по годам для выбранной профессии
                    statistics4 (dict): Динамика количества вакансий по годам для выбранной профессии
                    statistics5 (dict): Уровень зарплат по городам (в порядке убывания)
                    statistics6 (dict): Доля вакансий по городам (в порядке убывания)
        """
        self.workbook = Workbook()
        self.vacancyName = vacancyName
        self.statistics = statistics
        self.statistics2 = statistics2
        self.statistics3 = statistics3
        self.statistics4 = statistics4
        self.statistics5 = statistics5
        self.statistics6 = statistics6

    def get_excel(self):
        """Генерирует эксель файл и таблицы в нём с помощью библиотеки openpyxl
        """
        c = self.workbook.active
        # задаём название таблицы и атрибуиов
        c.title = 'Статистика по годам'
        c.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancyName, 'Количество вакансий',
                    'Количество вакансий - ' + self.vacancyName])
        for year in self.statistics.keys():
            c.append([year, self.statistics[year], self.statistics3[year], self.statistics2[year],
                      self.statistics4[year]])
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancyName, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancyName]]
        # в цикле находим необходимое количесвто столбцов
        columnWidth = []
        for x in data:
            for i, cell in enumerate(x):
                if len(columnWidth) > i:
                    if len(cell) > columnWidth[i]:
                        columnWidth[i] = len(cell)
                else:
                    columnWidth += [len(cell)]
        for i, columnWidth in enumerate(columnWidth, 1):
            # устанавливаем ширину столбцов для вместимости самой длинной строки в столбце
            c.column_dimensions[get_column_letter(i)].width = 2 + columnWidth
        newData = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]
        for (city, value), (city2, value2) in zip(self.statistics5.items(), self.statistics6.items()):
            newData.append([city, value, '', city2, value2])
        d = self.workbook.create_sheet('Статистика по городам')
        for x in newData:
            d.append(x)
        # в цикле находим необходимое количесвто столбцов
        columnWidth = []
        for x in newData:
            for i, cellule in enumerate(x):
                cellule = str(cellule)
                if len(columnWidth) > i:
                    if len(cellule) > columnWidth[i]:
                        columnWidth[i] = len(cellule)
                else:
                    columnWidth += [len(cellule)]
        for i, columnWidth in enumerate(columnWidth, 1):
            # устанавливаем ширину столбцов для вместимости самой длинной строки в столбце
            d.column_dimensions[get_column_letter(i)].width = columnWidth + 2
        fontBold = Font(bold=True)
        # заполняем таблицу данными
        for column in 'ABCDE':
            c[column + '1'].font = fontBold
            d[column + '1'].font = fontBold
        for index, _ in enumerate(self.statistics5):
            d['E' + str(2 + index)].number_format = '0.00%'
        fontThin = Side(border_style='thin', color='00000000')
        for x in range(len(newData)):
            for column in 'ABDE':
                d[column + str(1 + x)].border = Border(left=fontThin, bottom=fontThin, right=fontThin, top=fontThin)
        self.statistics[1] = 1
        for x, _ in enumerate(self.statistics):
            for column in 'ABCDE':
                c[column + str(1 + x)].border = Border(left=fontThin, bottom=fontThin, right=fontThin, top=fontThin)
        self.workbook.save(filename='report.xlsx') #сохраняем полученную excel таблицу

    def get_graphics(self):
        """Генерирует графики и диаграммы с помощью библиотек matplotlib, numpy
        """
        fig, ((axes, axes2), (axes3, axes4)) = plt.subplots(nrows=2, ncols=2)
        bar = axes.bar(np.array(list(self.statistics.keys())) - 0.4, self.statistics.values(), width=0.4)
        bar2 = axes.bar(np.array(list(self.statistics.keys())), self.statistics3.values(), width=0.4)
        #создаём первый вариант графика
        axes.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        axes.grid(axis='y')
        axes.legend((bar[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancyName.lower()), prop={'size': 8})
        axes.set_xticks(np.array(list(self.statistics.keys())) - 0.2, list(self.statistics.keys()), rotation=90)
        axes.xaxis.set_tick_params(labelsize=8)
        axes.yaxis.set_tick_params(labelsize=8)
        # создаём второй вариант графика
        axes2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar = axes2.bar(np.array(list(self.statistics2.keys())) - 0.4, self.statistics2.values(), width=0.4)
        bar2 = axes2.bar(np.array(list(self.statistics2.keys())), self.statistics4.values(), width=0.4)
        axes2.legend((bar[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancyName.lower()),
                     prop={'size': 8})
        axes2.set_xticks(np.array(list(self.statistics2.keys())) - 0.2, list(self.statistics2.keys()), rotation=90)
        axes2.grid(axis='y')
        axes2.xaxis.set_tick_params(labelsize=8)
        axes2.yaxis.set_tick_params(labelsize=8)
        # создаём третий вариант графика
        axes3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        axes3.barh(
            list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.statistics5.keys()))]),
            list(reversed(list(self.statistics5.values()))), color='blue', height=0.5, align='center')
        axes3.yaxis.set_tick_params(labelsize=6)
        axes3.xaxis.set_tick_params(labelsize=8)
        axes3.grid(axis='x')
        # создаём четвёртый вариант графика
        axes4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        e = 1 - sum([value for value in self.statistics6.values()])
        axes4.pie(list(self.statistics6.values()) + [e], labels=list(self.statistics6.keys()) + ['Другие'],
                  textprops={'fontsize': 6})
        #выводим все 4 картинки и сохраняе файл
        plt.tight_layout()
        plt.savefig('graph.png')

    #метод формирует pdf файл, содержащий в себе таблицу и диаграммы, с помощью библиотек pathlib, pdfkit, jinja2
    def get_pdf(self):
        """Генерирует pgf файл, содержащий таблицу и диаграммы, с помощью библиотек pathlib, pdfkit, jinja2
        """
        statistic = [] #создаём список для всех необходимых статистик
        environment = Environment(loader=FileSystemLoader('.'))
        template = environment.get_template("html.html")
        #заполняем статистики данными
        for i in self.statistics.keys():
            statistic.append([i, self.statistics[i], self.statistics2[i], self.statistics3[i],
                              self.statistics4[i]])
        for x in self.statistics6:
            self.statistics6[x] = round(100 * self.statistics6[x], 2)
        #на основе только что заполненных нами данных формируем pdf файл
        pdfTemp = template.render(
            {'name': self.vacancyName, 'path': '{0}/{1}'.format(pathlib.Path(__file__).parent.resolve(), 'graph.png'),
             'stats': statistic, 'stats5': self.statistics5, 'stats6': self.statistics6})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdfTemp, 'report.pdf', configuration=config, options={"enable-local-file-access": None})

class InputConnect:
    """Класс отвечает за обработку параметров вводимых пользователем (фильтры, сортировка, диапазон вывода, требуемые столбцы),
     а также за вывод данных на экран
            Attributes:
                fileName (str): Название файла
                vacancyName (str): Название профессии для статистики
        """
    def __init__(self):
        self.fileName = input('Введите название файла: ')
        self.vacancyName = input('Введите название профессии: ')
        a = DataSet(self.fileName, self.vacancyName)
        statistics, statistics2, statistics3, statistics4, statistics5, statistics6 = a.get_statistic()
        a.print_statistic(statistics, statistics2, statistics3, statistics4, statistics5, statistics6)
        b = Report(self.vacancyName, statistics, statistics2, statistics3, statistics4, statistics5, statistics6)
        b.get_excel()
        b.get_graphics()
        b.get_pdf()

def get_answer():
    """Запускает программу
    """
    InputConnect()